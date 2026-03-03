"""
Módulo agente de extracción de facturas.
- Conexión IMAP a Gmail
- Extracción de datos con Claude AI (Anthropic)
- Procesamiento de PDF, DOCX, XLSX adjuntos
- Guardado en Google Sheets
NO contiene rutas FastAPI — eso es responsabilidad de backend.py
"""

from dotenv import load_dotenv
load_dotenv()

import os
import json
import imaplib
import email
from email.header import decode_header
import io
import re
from datetime import datetime
from typing import Optional

# ── CONSTANTES ────────────────────────────────────────────────────────────────
_BASE_DIR            = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH           = os.getenv("EXCEL_OUTPUT_PATH", os.path.join(_BASE_DIR, "facturas_seguimiento.xlsx"))
PROCESSED_FILE       = os.path.join(_BASE_DIR, "processed_emails.json")
IMAP_HOST            = os.getenv("IMAP_HOST", "imap.gmail.com")
IMAP_PORT            = int(os.getenv("IMAP_PORT", "993"))
IMAP_FOLDER          = os.getenv("IMAP_FOLDER", "INBOX")
EMAIL_USER           = os.getenv("EMAIL_USER", "")
EMAIL_PASS           = os.getenv("EMAIL_PASS", "")
ANTHROPIC_API_KEY    = os.getenv("ANTHROPIC_API_KEY", "")
GOOGLE_SHEETS_ID     = os.getenv("GOOGLE_SHEETS_ID", "")
SERVICE_ACCOUNT_FILE = os.path.join(_BASE_DIR, os.getenv("SERVICE_ACCOUNT_FILE", "service_account.json"))
SHEET_NAME           = "Facturas"

# Columnas exactas en Google Sheets
COLUMNS = [
    "Mes", "Fecha Factura", "Número Factura", "Proveedor",
    "ID", "Número ID", "Subtotal", "Descuento", "IVA", "Rete IVA",
    "Rete ICA", "Impto Consumo", "Propina", "Otros Impuestos",
    "Retención en la fuente", "Administración", "Utilidad", "Imprevistos",
    "Valor Total", "Clasificación", "Estado", "Valor Pagado",
    "Valor Por Pagar", "Fecha Pago", "Cliente", "Cotización Inventto",
    "Observaciones", "Email Origen", "Fecha Procesado"
]

# ── GOOGLE SHEETS CLIENT ──────────────────────────────────────────────────────
_gspread_client = None


def get_sheet():
    """Retorna la hoja de Google Sheets, inicializando el cliente si es necesario."""
    global _gspread_client
    import gspread
    from google.oauth2.service_account import Credentials

    SCOPES = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]

    if _gspread_client is None:
        creds_json_env = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if creds_json_env:
            creds_dict = json.loads(creds_json_env)
            creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        else:
            creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        _gspread_client = gspread.authorize(creds)

    spreadsheet = _gspread_client.open_by_key(GOOGLE_SHEETS_ID)
    try:
        return spreadsheet.worksheet(SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=SHEET_NAME, rows=1000, cols=len(COLUMNS))
        ws.append_row(COLUMNS)
        return ws


def setup_sheets():
    """Crea encabezados en Google Sheets si la hoja está vacía."""
    try:
        ws = get_sheet()
        existing = ws.row_values(1)
        if not existing:
            ws.append_row(COLUMNS)
            print(f"✅ Google Sheets inicializado con {len(COLUMNS)} columnas")
        else:
            print(f"✅ Google Sheets conectado: {SHEET_NAME} ({len(existing)} columnas)")
    except Exception as e:
        print(f"❌ Error conectando a Google Sheets: {e}")
        raise


def _load_processed_ids() -> set:
    try:
        if os.path.exists(PROCESSED_FILE):
            with open(PROCESSED_FILE) as f:
                return set(json.load(f))
    except Exception:
        pass
    return set()


def _save_processed_ids(ids: set):
    try:
        with open(PROCESSED_FILE, "w") as f:
            json.dump(list(ids), f)
    except Exception as e:
        print(f"⚠️ No se pudo guardar IDs procesados: {e}")


# ── IMAP ──────────────────────────────────────────────────────────────────────
def connect_imap() -> imaplib.IMAP4_SSL:
    mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    mail.login(EMAIL_USER, EMAIL_PASS)
    mail.select(IMAP_FOLDER)
    print(f"✅ Conectado a {IMAP_HOST} como {EMAIL_USER}")
    return mail


def search_invoice_emails(mail: imaplib.IMAP4_SSL) -> list:
    """Devuelve los IDs de los 20 correos más recientes."""
    try:
        _, data = mail.search(None, "ALL")
        all_ids = [uid.decode() for uid in data[0].split()]
        recientes = all_ids[-20:]
        print(f"📬 {len(recientes)} correos para analizar")
        return recientes
    except Exception as e:
        print(f"❌ Error buscando correos: {e}")
        return []


# ── EXTRACCIÓN DE TEXTO ───────────────────────────────────────────────────────
def _decode_str(s) -> str:
    if not s:
        return ""
    parts = decode_header(s)
    result = []
    for part, enc in parts:
        if isinstance(part, bytes):
            result.append(part.decode(enc or "utf-8", errors="replace"))
        else:
            result.append(str(part))
    return " ".join(result)


def _extract_pdf_text(data: bytes) -> str:
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(data))
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception as e:
        return f"[Error PDF: {e}]"


def _extract_docx_text(data: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        return f"[Error DOCX: {e}]"


def _extract_xlsx_text(data: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                line = " | ".join(str(c) for c in row if c is not None)
                if line.strip():
                    lines.append(line)
        return "\n".join(lines[:100])
    except Exception as e:
        return f"[Error XLSX: {e}]"


def _get_email_content(msg) -> tuple:
    """Retorna (cuerpo_texto, lista_de_textos_adjuntos)."""
    body = ""
    attachments = []

    for part in msg.walk():
        ct = part.get_content_type()
        cd = str(part.get("Content-Disposition", ""))

        if "attachment" in cd or ct in (
            "application/pdf",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ):
            filename = part.get_filename() or ""
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            if filename.lower().endswith(".pdf") or ct == "application/pdf":
                attachments.append(_extract_pdf_text(payload))
            elif filename.lower().endswith(".docx"):
                attachments.append(_extract_docx_text(payload))
            elif filename.lower().endswith(".xlsx"):
                attachments.append(_extract_xlsx_text(payload))

        elif ct == "text/plain" and "attachment" not in cd:
            try:
                body += part.get_payload(decode=True).decode(
                    part.get_content_charset() or "utf-8", errors="replace"
                )
            except Exception:
                pass
        elif ct == "text/html" and not body and "attachment" not in cd:
            try:
                from bs4 import BeautifulSoup
                raw = part.get_payload(decode=True).decode(
                    part.get_content_charset() or "utf-8", errors="replace"
                )
                body += BeautifulSoup(raw, "html.parser").get_text(separator="\n")
            except Exception:
                pass

    return body, attachments


# ── INTELIGENCIA ARTIFICIAL (Claude) ─────────────────────────────────────────
def _extract_with_ai(subject: str, body: str, attachments: list) -> Optional[dict]:
    """
    Usa Claude (Anthropic) para extraer datos de la factura.
    Retorna dict con las claves de COLUMNS o None si no es factura.

    IMPORTANTE: el prompt NO usa f-string para la plantilla JSON,
    porque Python interpretaría {clave} como format specifiers y lanzaría
    "Invalid format specifier".  Solo se usa f-string para full_content.
    """
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        content_parts = [f"Asunto: {subject}", f"Cuerpo:\n{body[:4000]}"]
        for i, att in enumerate(attachments, 1):
            content_parts.append(f"Adjunto {i}:\n{att[:4000]}")
        full_content = "\n\n".join(content_parts)

        # ⚠️  String normal (NO f-string) para la plantilla JSON
        #    Si usáramos f-string, Python intentaría evaluar {Mes}, {Subtotal}, etc.
        prompt = (
            "Analiza el siguiente correo y determina si contiene una factura.\n\n"
            "Si NO es una factura, responde únicamente: NO_ES_FACTURA\n\n"
            "Si SÍ es una factura, responde SOLO con un objeto JSON válido, "
            "sin texto adicional ni bloques markdown. Usa exactamente estas claves "
            "(deja en blanco o en 0 lo que no encuentres):\n\n"
            '{\n'
            '  "Mes": "",\n'
            '  "Fecha Factura": "DD/MM/YYYY",\n'
            '  "Número Factura": "",\n'
            '  "Proveedor": "",\n'
            '  "ID": "",\n'
            '  "Número ID": "",\n'
            '  "Subtotal": 0,\n'
            '  "Descuento": 0,\n'
            '  "IVA": 0,\n'
            '  "Rete IVA": 0,\n'
            '  "Rete ICA": 0,\n'
            '  "Impto Consumo": 0,\n'
            '  "Propina": 0,\n'
            '  "Otros Impuestos": 0,\n'
            '  "Retención en la fuente": 0,\n'
            '  "Administración": 0,\n'
            '  "Utilidad": 0,\n'
            '  "Imprevistos": 0,\n'
            '  "Valor Total": 0,\n'
            '  "Clasificación": "",\n'
            '  "Estado": "PENDIENTE",\n'
            '  "Valor Pagado": 0,\n'
            '  "Valor Por Pagar": 0,\n'
            '  "Fecha Pago": "",\n'
            '  "Cliente": "",\n'
            '  "Cotización Inventto": "",\n'
            '  "Observaciones": ""\n'
            '}\n\n'
            "CORREO A ANALIZAR:\n"
            + full_content
        )

        resp = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}]
        )
        text = resp.content[0].text.strip()

        if "NO_ES_FACTURA" in text:
            return None

        # Parser de 3 niveles para robusted ante distintos formatos de respuesta
        def _parse_json(raw: str) -> Optional[dict]:
            # 1) Intentar parsear directamente
            try:
                return json.loads(raw)
            except Exception:
                pass
            # 2) Buscar bloque ```json ... ``` o ``` ... ```
            block = re.search(r'```(?:json)?\s*([\s\S]*?)```', raw)
            if block:
                try:
                    return json.loads(block.group(1).strip())
                except Exception:
                    pass
            # 3) Buscar el bloque JSON con llaves balanceadas
            start = raw.find('{')
            if start == -1:
                return None
            depth = 0
            for i, ch in enumerate(raw[start:], start):
                if ch == '{':
                    depth += 1
                elif ch == '}':
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(raw[start:i + 1])
                        except Exception:
                            return None
            return None

        data = _parse_json(text)
        if data:
            return data

        print(f"⚠️ Claude no devolvió JSON válido. Respuesta: {text[:200]}")
        return None

    except Exception as e:
        print(f"⚠️ Error IA: {e}")
        return None


def _is_duplicate_invoice(numero_factura: str) -> bool:
    """Comprueba si la factura ya existe en Google Sheets por Número Factura."""
    if not numero_factura or numero_factura in ("N/A", "0", ""):
        return False
    try:
        ws = get_sheet()
        headers = ws.row_values(1)
        # Buscar la columna "Número Factura"
        try:
            col_idx = headers.index("Número Factura") + 1
        except ValueError:
            col_idx = 3  # fallback: columna 3
        col_values = ws.col_values(col_idx)
        for val in col_values[1:]:  # saltar encabezado
            if str(val).strip() == str(numero_factura).strip():
                return True
        return False
    except Exception:
        return False


# ── GUARDAR EN GOOGLE SHEETS ──────────────────────────────────────────────────
def save_to_sheets(invoice_data: dict, email_from: str):
    """
    Agrega una fila a Google Sheets con los datos extraídos.
    El orden de columnas sigue exactamente COLUMNS.
    """
    try:
        numero    = str(invoice_data.get("Número Factura") or "N/A")
        proveedor = str(invoice_data.get("Proveedor") or "N/A")

        if _is_duplicate_invoice(numero):
            print(f"⏭️ Factura duplicada omitida: {numero} — {proveedor}")
            return

        def v(key, default=""):
            """Obtiene valor del dict con fallback."""
            val = invoice_data.get(key)
            return val if val is not None else default

        row = [
            v("Mes"),
            v("Fecha Factura"),
            numero,
            proveedor,
            v("ID"),
            v("Número ID"),
            v("Subtotal", 0),
            v("Descuento", 0),
            v("IVA", 0),
            v("Rete IVA", 0),
            v("Rete ICA", 0),
            v("Impto Consumo", 0),
            v("Propina", 0),
            v("Otros Impuestos", 0),
            v("Retención en la fuente", 0),
            v("Administración", 0),
            v("Utilidad", 0),
            v("Imprevistos", 0),
            v("Valor Total", 0),
            v("Clasificación"),
            v("Estado", "PENDIENTE"),
            v("Valor Pagado", 0),
            v("Valor Por Pagar", 0),
            v("Fecha Pago"),
            v("Cliente"),
            v("Cotización Inventto"),
            v("Observaciones"),
            email_from,
            datetime.now().strftime("%d/%m/%Y %H:%M"),
        ]

        ws = get_sheet()
        ws.append_row(row, value_input_option="USER_ENTERED")
        print(f"✅ Factura guardada en Sheets: {numero} — {proveedor}")

    except Exception as e:
        print(f"❌ Error guardando en Sheets: {e}")
        import traceback
        print(traceback.format_exc())


def export_to_excel() -> str:
    """Descarga todos los registros de Sheets y genera un Excel local (bajo demanda)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    try:
        ws = get_sheet()
        rows = ws.get_all_values()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Facturas"
        for i, row in enumerate(rows, 1):
            sheet.append(row)
            if i == 1:
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1B4FD8")
                    cell.alignment = Alignment(horizontal="center")
        wb.save(EXCEL_PATH)
        print(f"✅ Excel exportado: {EXCEL_PATH}")
        return EXCEL_PATH
    except Exception as e:
        print(f"❌ Error exportando Excel: {e}")
        raise


# ── PROCESO PRINCIPAL ──────────────────────────────────────────────────────────
def process_emails():
    """Función principal del agente. Conecta, lee, extrae y guarda."""
    if not EMAIL_USER or not EMAIL_PASS:
        print("❌ Credenciales de correo no configuradas en .env")
        return
    if not ANTHROPIC_API_KEY:
        print("❌ ANTHROPIC_API_KEY no configurada en .env")
        return

    setup_sheets()
    processed_ids = _load_processed_ids()

    print(f"📧 Conectando a Gmail ({EMAIL_USER})...")
    try:
        mail = connect_imap()
    except Exception as e:
        print(f"❌ Error de conexión IMAP: {e}")
        return

    email_ids = search_invoice_emails(mail)
    nuevos = [eid for eid in email_ids if eid not in processed_ids]
    print(f"🔍 {len(nuevos)} correos nuevos de {len(email_ids)} encontrados")

    facturas_encontradas = 0

    for eid in nuevos:
        try:
            _, data = mail.fetch(eid, "(RFC822)")
            msg = email.message_from_bytes(data[0][1])
            subject = _decode_str(msg.get("Subject", "(sin asunto)"))
            from_   = _decode_str(msg.get("From", ""))

            print(f"📨 Analizando: {subject[:60]}")

            body, attachments = _get_email_content(msg)
            invoice_data = _extract_with_ai(subject, body, attachments)

            if invoice_data:
                save_to_sheets(invoice_data, from_)
                facturas_encontradas += 1

            processed_ids.add(eid)
            _save_processed_ids(processed_ids)

        except Exception as e:
            print(f"⚠️ Error procesando correo {eid}: {e}")
            continue

    try:
        mail.logout()
    except Exception:
        pass

    print(f"✅ Proceso completado — {facturas_encontradas} factura(s) extraída(s) de {len(nuevos)} correos")


# ── ENTRY POINT STANDALONE ─────────────────────────────────────────────────────
if __name__ == "__main__":
    process_emails()
